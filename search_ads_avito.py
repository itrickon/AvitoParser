import os
import re
import asyncio
from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook
from deep_translator import GoogleTranslator


class SearchAvitoAds:
    def __init__(self, city, keyword, max_num_ads=10):
        self.city = city
        self.keyword = keyword
        self.max_num_ads = max_num_ads
        self.ads = []
        self.data_saving = "avito_parse_results/avito_ads.xlsx"
        self.warning_message()

    async def _get_links(self):
        link_selector = '[data-marker="item-title"][href]'
        found_links = await self.page.query_selector_all(link_selector)

        links = []
        for link in found_links:
            href = await link.get_attribute("href")
            links.append(f"https://www.avito.ru/{href}")

        return links

    async def _go_to_next_page(self):
        """Переход на следующую страницу"""
        try:
            # Ищем кнопку для перехода на след. страницу
            next_button = await self.page.query_selector('[aria-label="Следующая страница"]')
            if next_button and await next_button.is_visible():
                await next_button.click()
                await asyncio.sleep(2)
                return True
            return False
        except Exception as e:
            print(f"Ошибка при переходе на следующую страницу: {e}")
            return False

    def _create_xlsx(self):
        """Создание XLSX файла с заголовками"""
        # Создаем папку, если ее нет
        os.makedirs("avito_parse_results", exist_ok=True)

        # Создаем новую рабочую область
        wb = Workbook()
        ws = wb.active
        ws.title = "Avito Ads"

        # Добавляем заголовки
        headers = ["Ссылка на объявление", "ID"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Сохраняем файл
        wb.save(self.data_saving)
        print(f"Создан файл: {self.data_saving}")

    def _save_to_xlsx(self):
        """Сохранение данных в XLSX файл"""
        # Если файл не существует, создаем его
        if not os.path.exists(self.data_saving):
            self._create_xlsx()

        # Открываем существующий файл
        wb = load_workbook(self.data_saving)
        ws = wb.active

        # Определяем с какой строки начинать запись (последняя заполненная строка + 1)
        start_row = ws.max_row + 1 if ws.max_row > 1 else 2

        # Записываем данные
        for i, ad in enumerate(self.ads, start=start_row):
            ws.cell(row=i, column=1, value=ad)  # Ссылка
            ws.cell(row=i, column=2, value=i - 1)  # ID (начинаем с 1)

        # Сохраняем файл
        wb.save(self.data_saving)
        print(f"Данные сохранены в файл: {self.data_saving}")

    async def translate_text(self, city):
        """Переводим город на английский"""
        # Проверяем, является ли слово английским (только латинские буквы)
        is_english = bool(re.match(r"^[a-zA-Z\s\-]+$", city))

        if is_english:
            # Если уже английское слово, просто форматируем
            city_clean = "-".join(city.split())
            return city_clean.lower()
        else:
            # Если русское слово - переводим
            translator = GoogleTranslator(source="ru", target="en")
            try:
                a = await asyncio.to_thread(translator.translate, city)
                a = "-".join(a.split())
                return a.lower()
            except Exception as e:
                print(f"Ошибка перевода: {e}")

    def warning_message(self):
        print("\n" + "=" * 50)
        print("EDUCATIONAL USE ONLY - NO WARRANTY PROVIDED")
        print("This parser may violate Terms of Service.")
        print("Use only for learning web scraping techniques.")
        print("Author not responsible for any legal consequences.")
        print("=" * 50 + "\n")

    async def parse_main(self, update_callback=None):
        # Создаем XLSX файл перед началом парсинга
        self._create_xlsx()
        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(headless=False)
            self.context = await browser.new_context()
            self.page = await self.context.new_page()

            trans_text = await self.translate_text(self.city)
            # Формируем URL с городом
            await self.page.goto(
                f"https://www.avito.ru/{trans_text}?cd=1&q={self.keyword}",
                wait_until="domcontentloaded",
            )

            # Ждем появление слова "Найти" на странице
            await self.page.wait_for_selector('text="Найти"', timeout=60000)

            # Собираем ссылки с нескольких страниц
            while len(self.ads) < self.max_num_ads:
                # Получаем ссылки с текущей страницы
                page_links = await self._get_links()

                # Добавляем новые ссылки
                for link in page_links:
                    if len(self.ads) < self.max_num_ads:
                        self.ads.append(link)
                        await asyncio.sleep(0.1)

                print(f"Всего собрано ссылок: {len(self.ads)} из {self.max_num_ads}")

                # Проверяем, нужно ли собирать еще ссылки
                if len(self.ads) >= self.max_num_ads:
                    print(f"Достигнуто необходимое количество ссылок: {self.max_num_ads}")
                    break

                # Пытаемся перейти на следующую страницу
                if not await self._go_to_next_page():
                    print("Больше нет страниц для парсинга")
                    break

                await asyncio.sleep(2)  # Асинхронная задержка между страницами

            # Финальное сохранение всех данных
            self._save_to_xlsx()

            # Выводим результат
            print(f"Количество строк в файле: {len(self.ads)}")

            await browser.close()


async def main():
    parser = SearchAvitoAds(city="Липецк", keyword="Игровая клавиатура", max_num_ads=5)
    await parser.parse_main()


if __name__ == "__main__":
    asyncio.run(main())
